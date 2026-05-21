from excel import (
    MatchResult,
    tsv_id_code_title,
    tsv_release_date,
    xlsx_bytes,
)


def make_result(id="1842", code="M001", title="어벤져스 엔드게임", date="2019-04-24"):
    return MatchResult(id=id, code=code, title=title, release_date=date)


def test_tsv_id_code_title_single_row():
    results = [make_result()]
    assert tsv_id_code_title(results) == "1842\tM001\t어벤져스 엔드게임"


def test_tsv_id_code_title_multiple_rows():
    results = [
        make_result(id="1", code="A", title="x"),
        make_result(id="2", code="B", title="y"),
    ]
    assert tsv_id_code_title(results) == "1\tA\tx\n2\tB\ty"


def test_tsv_id_code_title_empty():
    assert tsv_id_code_title([]) == ""


def test_tsv_release_date_single_row():
    assert tsv_release_date([make_result(date="2019-04-24")]) == "2019-04-24"


def test_tsv_release_date_multiple_rows():
    results = [
        make_result(date="2019-04-24"),
        make_result(date="2024-06-12"),
    ]
    assert tsv_release_date(results) == "2019-04-24\n2024-06-12"


def test_tsv_release_date_empty_dates_become_blank_lines():
    results = [
        make_result(date="2019-04-24"),
        make_result(date=""),
    ]
    assert tsv_release_date(results) == "2019-04-24\n"


def test_xlsx_bytes_produces_valid_workbook():
    from openpyxl import load_workbook
    import io
    blob = xlsx_bytes([make_result()])
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 첫 행은 헤더
    assert rows[0] == ("id", "code", "title", "개봉일")
    assert rows[1] == ("1842", "M001", "어벤져스 엔드게임", "2019-04-24")
